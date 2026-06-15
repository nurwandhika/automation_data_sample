import os
import glob
import csv
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from collections import defaultdict

DATE_FORMATS = [
    "%Y-%m-%d",
    "%d-%m-%Y",
    "%d/%m/%Y",
    "%m/%d/%Y",
    "%Y/%m/%d",
    "%d-%b-%Y",
    "%d %b %Y",
]


def normalize_text(value):
    return re.sub(r"[^a-z0-9]+", " ", str(value).lower()).strip()


def parse_date_value(value):
    text = str(value).strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def build_rule_definitions(match_rules, headers, side):
    definitions = []
    for rule in match_rules:
        column_name = rule[f"{side}_column"]
        column_index = headers.index(column_name)
        definitions.append({
            "column": column_name,
            "index": column_index,
            "match_mode": rule.get("match_mode", "exact"),
            "tolerance": int(rule.get("tolerance", 0) or 0),
            "value_type": rule.get("value_type", "text"),
        })
    return definitions


def normalize_contains_mode(match_mode, side):
    if match_mode == "contains":
        return "core_contains_swi" if side == "core" else "swi_contains_core"
    return match_mode


def normalize_text_rule_mode(match_mode, side):
    if match_mode == "contains":
        return "core_contains_swi" if side == "core" else "swi_contains_core"
    if match_mode == "starts_with":
        return "core_starts_with_swi" if side == "core" else "swi_starts_with_core"
    if match_mode == "ends_with":
        return "core_ends_with_swi" if side == "core" else "swi_ends_with_core"
    return match_mode


def text_contains_match(container_value, contained_value):
    container_text = normalize_text(container_value)
    contained_text = normalize_text(contained_value)

    if not container_text or not contained_text:
        return container_text == contained_text

    return contained_text in container_text


def build_value_signature(raw_value, value_type, match_mode, tolerance):
    text = str(raw_value).strip()
    if text == "":
        return "__blank__"

    if match_mode in {"core_contains_swi", "swi_contains_core"}:
        return normalize_text(text)

    if match_mode in {"core_starts_with_swi", "swi_starts_with_core", "core_ends_with_swi", "swi_ends_with_core"}:
        return normalize_text(text)

    if match_mode == "integer_tolerance":
        number_match = re.fullmatch(r"[-+]?\d+", text)
        if not number_match:
            return f"int:{normalize_text(text)}"
        number = int(text)
        interval = max(tolerance, 1)
        return f"int:{number // interval}"

    if match_mode == "date_tolerance":
        parsed_date = parse_date_value(text)
        if parsed_date is None:
            return f"date:{normalize_text(text)}"
        return f"date:{parsed_date.toordinal()}"

    if value_type == "integer":
        number_match = re.fullmatch(r"[-+]?\d+", text)
        if number_match:
            return f"int:{int(text)}"
        return f"int:{normalize_text(text)}"

    if value_type == "date":
        parsed_date = parse_date_value(text)
        if parsed_date is not None:
            return f"date:{parsed_date.toordinal()}"
        return f"date:{normalize_text(text)}"

    return normalize_text(text)


def build_row_signature(row, rule_definitions):
    signature_parts = []
    for rule in rule_definitions:
        raw_value = row[rule["index"]] if len(row) > rule["index"] else ""
        signature_parts.append(
            build_value_signature(
                raw_value=raw_value,
                value_type=rule["value_type"],
                match_mode=rule["match_mode"],
                tolerance=rule["tolerance"],
            )
        )
    return tuple(signature_parts)


def rows_match(core_row, swi_row, core_rule_definitions, swi_rule_definitions):
    canonical_parts = []

    for core_rule, swi_rule in zip(core_rule_definitions, swi_rule_definitions):
        core_value = core_row[core_rule["index"]] if len(core_row) > core_rule["index"] else ""
        swi_value = swi_row[swi_rule["index"]] if len(swi_row) > swi_rule["index"] else ""

        core_mode = normalize_text_rule_mode(core_rule["match_mode"], "core")
        swi_mode = normalize_text_rule_mode(swi_rule["match_mode"], "swi")

        if core_mode == "date_tolerance" or swi_mode == "date_tolerance":
            core_date = parse_date_value(core_value)
            swi_date = parse_date_value(swi_value)
            if core_date is None or swi_date is None:
                if normalize_text(core_value) != normalize_text(swi_value):
                    return None
                canonical_parts.append(("date_tolerance", normalize_text(core_value)))
                continue

            tolerance_days = max(core_rule["tolerance"], swi_rule["tolerance"])
            if abs((core_date - swi_date).days) > tolerance_days:
                return None

            canonical_parts.append(("date_tolerance", f"{core_date.toordinal()}:{swi_date.toordinal()}:{tolerance_days}"))
            continue

        if core_mode in {"core_starts_with_swi", "swi_starts_with_core", "core_ends_with_swi", "swi_ends_with_core"} or swi_mode in {"core_starts_with_swi", "swi_starts_with_core", "core_ends_with_swi", "swi_ends_with_core"}:
            if core_mode == "core_starts_with_swi" or swi_mode == "swi_starts_with_core":
                if not normalize_text(core_value).startswith(normalize_text(swi_value)):
                    return None
                canonical_parts.append(("starts_with", normalize_text(str(swi_value))))
                continue

            if core_mode == "swi_starts_with_core" or swi_mode == "core_starts_with_swi":
                if not normalize_text(swi_value).startswith(normalize_text(core_value)):
                    return None
                canonical_parts.append(("starts_with", normalize_text(str(core_value))))
                continue

            if core_mode == "core_ends_with_swi" or swi_mode == "swi_ends_with_core":
                if not normalize_text(core_value).endswith(normalize_text(swi_value)):
                    return None
                canonical_parts.append(("ends_with", normalize_text(str(swi_value))))
                continue

            if core_mode == "swi_ends_with_core" or swi_mode == "core_ends_with_swi":
                if not normalize_text(swi_value).endswith(normalize_text(core_value)):
                    return None
                canonical_parts.append(("ends_with", normalize_text(str(core_value))))
                continue

        if core_mode in {"core_contains_swi", "swi_contains_core"} or swi_mode in {"core_contains_swi", "swi_contains_core"}:
            if core_mode == "core_contains_swi" or swi_mode == "swi_contains_core":
                if not text_contains_match(core_value, swi_value):
                    return None
                canonical_parts.append(("contains", normalize_text(str(swi_value))))
                continue

            if core_mode == "swi_contains_core" or swi_mode == "core_contains_swi":
                if not text_contains_match(swi_value, core_value):
                    return None
                canonical_parts.append(("contains", normalize_text(str(core_value))))
                continue

        core_signature = build_value_signature(core_value, core_rule["value_type"], core_mode, core_rule["tolerance"])
        swi_signature = build_value_signature(swi_value, swi_rule["value_type"], swi_mode, swi_rule["tolerance"])

        if core_signature != swi_signature:
            return None

        canonical_parts.append(("match", core_signature))

    return tuple(canonical_parts)


def dynamic_vertical_reconciliation(folder_a, folder_b, match_col_core=None, match_col_swi=None, output_file=None, output_format="split_tables", match_rules=None):
    print(f"Scanning '{folder_a}' and '{folder_b}' for CSV files...")
    
    # 1. Dynamically find the files
    file_a_list = glob.glob(os.path.join(folder_a, "*.csv"))
    file_b_list = glob.glob(os.path.join(folder_b, "*.csv"))
    
    if len(file_a_list) != 1 or len(file_b_list) != 1:
        print(f"Error: Please place exactly ONE CSV file in '{folder_a}' and ONE CSV file in '{folder_b}'.")
        return

    # Extract exact filenames
    core_file_path = file_a_list[0]
    swi_file_path = file_b_list[0]
    
    core_filename = os.path.basename(core_file_path).replace('.csv', '')
    swi_filename = os.path.basename(swi_file_path).replace('.csv', '')

    print(f"Loaded File A: {core_filename}")
    print(f"Loaded File B: {swi_filename}")
    
    # 2. Read raw CSV data
    with open(core_file_path, mode='r', encoding='utf-8-sig') as f:
        core_data = list(csv.reader(f))
    
    with open(swi_file_path, mode='r', encoding='utf-8-sig') as f:
        swi_data = list(csv.reader(f))

    core_headers = core_data[0]
    swi_headers = swi_data[0]

    if match_rules:
        core_rule_definitions = build_rule_definitions(match_rules, core_headers, "core")
        swi_rule_definitions = build_rule_definitions(match_rules, swi_headers, "swi")
        uses_contains = any(
            rule.get("match_mode") in {"contains", "core_contains_swi", "swi_contains_core", "starts_with", "ends_with", "core_starts_with_swi", "swi_starts_with_core", "core_ends_with_swi", "swi_ends_with_core"}
            for rule in match_rules
        )
        uses_date_tolerance = any(rule.get("match_mode") == "date_tolerance" for rule in match_rules)

        if uses_contains or uses_date_tolerance:
            core_map = defaultdict(list)
            swi_map = defaultdict(list)
            pair_map = defaultdict(lambda: {"c_rows": [], "s_rows": [], "c_seen": set(), "s_seen": set()})
            matched_core_rows = set()
            matched_swi_rows = set()

            for core_index, core_row in enumerate(core_data[1:], start=1):
                for swi_index, swi_row in enumerate(swi_data[1:], start=1):
                    pair_key = rows_match(core_row, swi_row, core_rule_definitions, swi_rule_definitions)
                    if pair_key is None:
                        continue

                    bucket = pair_map[pair_key]
                    core_key = tuple(core_row)
                    swi_key = tuple(swi_row)

                    if core_key not in bucket["c_seen"]:
                        bucket["c_rows"].append(core_row)
                        bucket["c_seen"].add(core_key)

                    if swi_key not in bucket["s_seen"]:
                        bucket["s_rows"].append(swi_row)
                        bucket["s_seen"].add(swi_key)

                    matched_core_rows.add(core_index)
                    matched_swi_rows.add(swi_index)

            for core_index, core_row in enumerate(core_data[1:], start=1):
                if core_index in matched_core_rows:
                    continue

                bucket = pair_map[("core_unmatched", core_index)]
                core_key = tuple(core_row)
                if core_key not in bucket["c_seen"]:
                    bucket["c_rows"].append(core_row)
                    bucket["c_seen"].add(core_key)

            for swi_index, swi_row in enumerate(swi_data[1:], start=1):
                if swi_index in matched_swi_rows:
                    continue

                bucket = pair_map[("swi_unmatched", swi_index)]
                swi_key = tuple(swi_row)
                if swi_key not in bucket["s_seen"]:
                    bucket["s_rows"].append(swi_row)
                    bucket["s_seen"].add(swi_key)

            for key, bucket in pair_map.items():
                core_map[key] = bucket["c_rows"]
                swi_map[key] = bucket["s_rows"]
        else:
            core_map = defaultdict(list)
            for row in core_data[1:]:
                core_map[build_row_signature(row, core_rule_definitions)].append(row)

            swi_map = defaultdict(list)
            for row in swi_data[1:]:
                swi_map[build_row_signature(row, swi_rule_definitions)].append(row)
    else:
        if match_col_core is None or match_col_swi is None:
            raise ValueError("Either match_rules or match_col_core/match_col_swi must be provided.")

        try:
            core_idx = core_headers.index(match_col_core)
            swi_idx = swi_headers.index(match_col_swi)
        except ValueError:
            print(f"Error: Could not find '{match_col_core}' or '{match_col_swi}' in the headers.")
            return

        # 3. Map the data by Reference Number
        core_map = defaultdict(list)
        for row in core_data[1:]:
            if len(row) > core_idx:
                core_map[normalize_text(row[core_idx])].append(row)

        swi_map = defaultdict(list)
        for row in swi_data[1:]:
            if len(row) > swi_idx:
                swi_map[normalize_text(row[swi_idx])].append(row)

    print(f"Executing {'text/date-aware' if match_rules and any(rule.get('match_mode') in {'contains', 'core_contains_swi', 'swi_contains_core', 'starts_with', 'ends_with', 'core_starts_with_swi', 'swi_starts_with_core', 'core_ends_with_swi', 'swi_ends_with_core', 'date_tolerance'} for rule in match_rules) else 'exact matching'} logic... (Format: {output_format})")
    
    # 4. Setup the Excel Output
    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliation"

    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    bold_font = Font(bold=True)
    
    # Pre-calculate all pairs and assign them an ID so colors sync perfectly across tables
    all_refs = set(core_map.keys()).union(set(swi_map.keys()))
    
    pairs_data = []
    pair_counter = 1
    
    for ref in sorted(all_refs, key=lambda value: str(value)):
        pairs_data.append({
            'pair_id': pair_counter,
            'c_rows': core_map.get(ref, []),
            's_rows': swi_map.get(ref, [])
        })
        pair_counter += 1

    # ==========================================
    # FORMAT 1: INTERLEAVED (Stacked row by row)
    # ==========================================
    if output_format == "interleaved":
        ws.append(["No"] + core_headers + ["file_name"])
        ws.append(["No"] + swi_headers + ["file_name"])
        
        for pair in pairs_data:
            c_rows, s_rows, pid = pair['c_rows'], pair['s_rows'], pair['pair_id']
            use_grey = (pid % 2 != 0) # Odd pairs are grey, Even pairs are white
            
            if c_rows and s_rows:
                max_len = max(len(c_rows), len(s_rows))
                for i in range(max_len):
                    if i < len(c_rows):
                        ws.append([pid] + c_rows[i] + [core_filename])
                        if use_grey:
                            for cell in ws[ws.max_row]: cell.fill = grey_fill
                    if i < len(s_rows):
                        ws.append([pid] + s_rows[i] + [swi_filename])
                        if use_grey:
                            for cell in ws[ws.max_row]: cell.fill = grey_fill
            elif c_rows:
                for row in c_rows:
                    ws.append([pid] + row + [core_filename])
                    for cell in ws[ws.max_row]: cell.fill = red_fill
            elif s_rows:
                for row in s_rows:
                    ws.append([pid] + row + [swi_filename])
                    for cell in ws[ws.max_row]: cell.fill = red_fill

    # ==========================================
    # FORMAT 2: SPLIT TABLES (Top and Bottom)
    # ==========================================
    elif output_format == "split_tables":
        # --- TABLE A (Core) ---
        ws.append([f"--- TABLE A: {core_filename} ---"])
        ws[f"A{ws.max_row}"].font = bold_font
        ws.append(["No"] + core_headers)
        ws[f"A{ws.max_row}"].font = bold_font
        
        for pair in pairs_data:
            c_rows, s_rows, pid = pair['c_rows'], pair['s_rows'], pair['pair_id']
            if not c_rows: continue # Skip if this pair has no data in File A
                
            use_grey = (pid % 2 != 0) # Base color on the exact Pair ID
            is_unmatched = not s_rows
            
            for row in c_rows:
                ws.append([pid] + row)
                if is_unmatched:
                    for cell in ws[ws.max_row]: cell.fill = red_fill
                elif use_grey:
                    for cell in ws[ws.max_row]: cell.fill = grey_fill

        # Visual Gap
        ws.append([]) 
        ws.append([])
        
        # --- TABLE B (Switcher) ---
        ws.append([f"--- TABLE B: {swi_filename} ---"])
        ws[f"A{ws.max_row}"].font = bold_font
        ws.append(["No"] + swi_headers)
        ws[f"A{ws.max_row}"].font = bold_font
        
        for pair in pairs_data:
            c_rows, s_rows, pid = pair['c_rows'], pair['s_rows'], pair['pair_id']
            if not s_rows: continue # Skip if this pair has no data in File B
                
            use_grey = (pid % 2 != 0) # Base color on the exact Pair ID
            is_unmatched = not c_rows
            
            for row in s_rows:
                ws.append([pid] + row)
                if is_unmatched:
                    for cell in ws[ws.max_row]: cell.fill = red_fill
                elif use_grey:
                    for cell in ws[ws.max_row]: cell.fill = grey_fill

    wb.save(output_file)
    print(f"Success! Output saved to: {output_file}")

# --- Trigger the Script ---
if __name__ == "__main__":
    # Ensure the directories exist
    folders_ready = True
    for folder in ['folder_a', 'folder_b']:
        if not os.path.exists(folder):
            os.makedirs(folder)
            print(f"Created '{folder}'.")
            folders_ready = False
            
    if not folders_ready:
        print("Please place your core (Acuan) CSV in 'folder_a' and switcher (Vendor) CSV in 'folder_b', then run again.")
    else:
        dynamic_vertical_reconciliation(
            folder_a='folder_a',
            folder_b='folder_b',
            match_col_core='no_ref',
            match_col_swi='no_ref',
            output_file='Automated_Grouped_Reconciliation.xlsx',
            output_format='split_tables' # Change to 'interleaved' for the other style!
        )